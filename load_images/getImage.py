import requests
import pandas as pd
import datetime
import os
import sys
from openpyxl import load_workbook
work_dir = os.path.dirname(__file__)
if getattr(sys, 'frozen', False):
    work_dir = os.path.dirname(sys.executable)
def download_image(url):
    if not url:
        print("URL不能为空")
        return
    image_name = ""
    if 'ExportFile' in url:
        parts = url.split('/')
        if len(parts) > 0:
            image_name = parts[-1]
    save_path = os.path.join(work_dir,'images', image_name)
    os.makedirs(os.path.dirname(work_dir), exist_ok=True)
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, "wb") as f:
            f.write(response.content)
        print(f"图片已成功下载到 {save_path}")
    else:
        print("请求失败，状态码：", response.status_code)

# 示例使用
# image_url = "https://cdn.customily.com/ExportFile/Benson/dd45ee2d-7876-41d8-a289-1e003b87593c.png"  # 替换为实际的图片URL
# download_image(image_url)
def get_current_time():
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S")
def get_excel_file(file_name,keep_default_na=False,dtype=str):
    try:
        file_path = os.path.join(work_dir, file_name)
        workbook = load_workbook(filename=file_path)
        sheet_names = workbook.sheetnames
        return pd.read_excel(file_path, sheet_name=sheet_names[0], keep_default_na=keep_default_na, dtype=dtype)
    except Exception as e:
        print(f"读取表格失败{file_name},{e}")
        return None
def main(): 
    excle = get_excel_file("Queue.xlsx")
    print("开始下载图片...")
    if excle is None:
        print("读取数据失败==>无法下载图片")
        return None
    excle.columns = [i.strip() for i in excle.columns]
    for index, row in excle.iterrows():
        url = row['图片链接']
        download_image(url)
    print("下载图片完成...")
#生成队表格式
if __name__ == "__main__":
    main()

