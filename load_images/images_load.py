import requests
import pandas as pd
import datetime
import os
import sys
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
work_dir = os.path.dirname(__file__)
if getattr(sys, 'frozen', False):
    work_dir = os.path.dirname(sys.executable)
def get_current_time():
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S")
# 下载图片
def download_image(url, index):
    tt = get_current_time()
    if not url:
        print("URL不能为空")
        return
    image_name = ""
    if 'ExportFile' in url:
        parts = url.split('/')
        if len(parts) > 0:
            image_name = parts[-1]
    save_path = os.path.join(work_dir, 'images', image_name)
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    response = requests.get(url)
    if response.status_code == 200:
        with open(save_path, "wb") as f:
            f.write(response.content)
        print(f"[{tt}]:第{index + 1}张图片已成功下载到 {save_path}")
    else:
        print(f"{tt}]:第{index + 1}张图片下载失败，状态码：{response.status_code}, URL:{url}")

def get_excel_file(file_name, keep_default_na=False, dtype=str):
    try:
        file_path = os.path.join(work_dir, file_name)
        workbook = load_workbook(filename=file_path)
        sheet_names = workbook.sheetnames
        return pd.read_excel(file_path, sheet_name=sheet_names[0], keep_default_na=keep_default_na, dtype=dtype)
    except Exception as e:
        print(f"读取表格失败{file_name},{e}")
        return None

def main():
    excle = get_excel_file("Queue.xlsx")
    print("开始下载图片...")
    if excle is None:
        print("读取数据失败==>无法下载图片")
        return None
    excle.columns = [i.strip() for i in excle.columns]
    max_workers = 100
    # 使用线程池
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        for index, row in excle.iterrows():
            try:
                url = row['图片链接'].strip()
                executor.submit(download_image, url, index)
            except Exception as e:
                print(f"第{index + 1}张图片下载失败,{url},{e}")

if __name__ == "__main__":
    main()