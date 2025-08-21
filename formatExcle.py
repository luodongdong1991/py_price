import pandas as pd
import datetime
import os
from openpyxl import load_workbook
# 获取当前工作目录
# work_dir = os.path.dirname(__file__)
work_dir = r"D:\实在RPA\抠图排查"
# 定义表格名称
sheet_name = "Recovered_Sheet1"
# 运行时间
current_time = datetime.datetime.now().strftime('%Y_%m_%d')
current_hour = datetime.datetime.now().hour
flag = 3
if 0 <= current_hour < 12:
    flag = 1
# 表头
ex_tran_no = "交易编号"
ex_url = "定制图片"
ex_zh_name = "商品中文名称"
ex_sort_no = "序号"
# if getattr(sys, 'frozen', False):
#     work_dir = os.path.dirname(sys.executable)
def get_excel_file(file_name,sheet_name,keep_default_na=False,dtype=str):
    try:
        file_path = os.path.join(work_dir, file_name)
        workbook = load_workbook(filename=file_path)
        sheet_names = workbook.sheetnames
        print(f"==表{sheet_names},共有{len(sheet_names)}个表格,{file_path}")
        return pd.read_excel(file_path, sheet_name=sheet_names[0], keep_default_na=keep_default_na, dtype=dtype)
    except Exception as e:
        print(f"读取表格失败{file_name},{e}")
        return None

def main():
    excle = get_excel_file("订单列表.xlsx", sheet_name)
    print("开始格式化数据...")
    if excle is None:
        print("读取数据失败==>无法格式化数据")
        return None
    excle.columns = [i.strip() for i in excle.columns]
    # 空 交易编号 处理
    row_list = excle.index[excle["交易编号"] == ""].tolist()
    if len(row_list) > 0:
        for i in row_list:
            index = i - 1
            excle.loc[[i], ex_tran_no] = excle.loc[index, ex_tran_no]
    duplicates = excle[excle.duplicated(subset=[ex_tran_no], keep=False)]
    if len(duplicates) > 0:
        for index, row in duplicates.iterrows():
            filtered_rows = excle[excle[ex_tran_no] == row[ex_tran_no]]
            if len(filtered_rows) > 0:
                count = -1
                for index, row in filtered_rows.iterrows():
                    count = 1 + count
                    if count >= 1:
                        excle.loc[[index], ex_tran_no] = excle.loc[index, ex_tran_no] +"-"+ str(count)
    # 过滤包含抠图的商品
    contains_koutu = excle[ex_zh_name].str.contains('抠图', na=False)
    excle = excle[contains_koutu]
    # 处理链接
    excle = excle[excle[ex_url].str.contains('ExportFile', na=False)]
    for index, row in excle.iterrows():
        url = row[ex_url]
        list = []
        if 'ExportFile' in url and ';' in url:
            list = url.split(';')
            if len(list) > 0:
                for item in list:
                    if 'ExportFile' in item:
                        excle.loc[[index], ex_url] = item
                        excle.loc[[index], ex_sort_no] = current_time + "_" + str(flag)
    # save
    print(excle.columns.tolist(),"原始表头")
    old_header = excle.columns.tolist()
    for i in range(len(old_header)):
        if  old_header[i] == "定制图片":
            old_header[i] = "图片链接" 
    excle.columns = old_header
    print(excle.columns.tolist(),"新")
    # 选取特定列
    selected_columns = excle[['交易编号', '图片链接',ex_sort_no]]
    print("开始保存数据...")
    name = f"Queue.xlsx"
    file_path = os.path.join(work_dir,name)
    selected_columns.to_excel(file_path,sheet_name="Sheet1",index=False,engine='openpyxl')
    print("保存成功")
    return True
#生成队表格式
if __name__ == "__main__":
    main()